"""This module provides a Table Type for Microsft Excel (.xlsx) files."""

import sys
from typing import TYPE_CHECKING, Any, List, Tuple, Union

from openpyxl import Workbook, load_workbook

from .basetabletype import BaseTableType

if TYPE_CHECKING:
    from pathlib import Path

    from tabler.table import Table


class XLSX(BaseTableType):
    """Table Type for Microsft Excel (.xlsx) files.

    :param str extension: Extension of file to save. Default .xlsx.
    :param verbose: If True print status messages. If None use
        :class:`tabler.tabletype.BaseTableType`.verbose.
    :type verbose: bool or None.
    """

    extensions: List[str] = [".xlsx"]
    empty_value: Any = None

    def __init__(self, extension: str = ".xlsx", verbose: bool = True):
        """Consturct :class:`tabler.tabletypes.XLSX`.

        :param str extension: Extension of file to save. Default .xlsx.
        :param verbose: If True print status messages. If None use
            :class:`tabler.tabletype.BaseTableType`.verbose.
        :type verbose: bool or None.
        """
        super().__init__(extension, verbose=verbose)

    def open_path(self, path: Union[str, "Path"]) -> Tuple[List[str], List[List[Any]]]:
        """Return header and rows from file.

        :param path: Path to file to be opened.
        :type path: str, pathlib.Path or compatible.
        """
        workbook = load_workbook(filename=str(path), read_only=True)
        worksheet = workbook.active
        data = [[cell.value for cell in row] for row in worksheet]
        workbook.close()
        return self.parse_row_data(data)

    def write(self, table: "Table", path: Union[str, "Path"]) -> None:
        """Save data from :class:`tabler.Table` to file.

        :param table:"Table" to save.
        :type table: :class:`tabler.Table`
        :param path: Path to file to be opened.
        :type path: str, pathlib.Path or compatible.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(table.header)
        for row in table:
            worksheet.append(list(row))
        workbook.save(str(path))
        print(
            "Written {} rows to file {}".format(len(table.rows), path), file=sys.stderr
        )
