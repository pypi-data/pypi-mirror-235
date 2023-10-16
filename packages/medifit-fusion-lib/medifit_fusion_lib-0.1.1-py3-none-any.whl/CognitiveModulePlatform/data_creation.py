import pandas as pd
import os
import openpyxl

class ExcelDataCreation:
    def __init__(self, file_path, sheets, rows):
        self.file_path = file_path
        self.sheets = sheets
        self.rows = rows

    def divide_excel_data(self, selected_output_path=None, unselected_output_path=None):
        dfs = pd.read_excel(self.file_path, sheet_name=self.sheets)
        if selected_output_path is None:
            selected_output_path = os.path.dirname(self.file_path)
        if unselected_output_path is None:
            unselected_output_path = os.path.dirname(self.file_path)
        self._process_and_save(dfs, selected_output_path, unselected_output_path)

    def _process_and_save(self, dfs, selected_output_path, unselected_output_path):
        selected_data = []
        unselected_data = []
        for sheet in self.sheets:
            df = dfs[sheet]
            selected = df.iloc[self.rows[sheet]].reset_index(drop=True)
            unselected = df[~df.index.isin(self.rows[sheet])].reset_index(drop=True)
            selected_data.append(selected)
            unselected_data.append(unselected)
        self._save_to_excel(selected_data, selected_output_path)
        self._save_to_excel(unselected_data, unselected_output_path)

    def _save_to_excel(self, data, output_path):
        merged_data = pd.concat(data, axis=0, keys=self.sheets, names=['sheet'])
        writer = pd.ExcelWriter(output_path, engine='xlsxwriter')
        for sheet in self.sheets:
            merged_data.loc[sheet].to_excel(writer, sheet_name=sheet, index=False, header=False)
        writer.save()

    def create_new_file(self, sheet, value_0, value_1):
        wb = openpyxl.load_workbook(self.file_path)
        sheet_to_copy = wb[sheet]
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        for row in sheet_to_copy.iter_rows(values_only=True):
            new_row = [value_0] + list(row)
            new_sheet.append(new_row)
        for s in value_1:
            other_sheet = wb[s]
            for row in other_sheet.iter_rows(values_only=True):
                new_row = [1] + list(row)
                new_sheet.append(new_row)
        new_wb.save(f'{os.path.dirname(self.file_path)}/{sheet}_new.xlsx')

# Usage
# Define the list of sheets
sheets = ['pf', 'th', 'pi', 'es', 'gr', 'mt', 'tn', 'tr', 'ad']

# Define the rows to select for each sheet
rows = {
    'pf': list(range(0,36)) + list(range(51, 86)) + list(range(101, 136)) + list(range(151, 191)) + list(range(206, 281)),
    'th': list(range(0, 101)) + list(range(146, 256)) + list(range(301, 366)),
    'pi': list(range(0, 36)) + list(range(51, 86)),
    'es': list(range(0, 36)) + list(range(51, 111)),
    'gr': list(range(0, 36)) + list(range(51, 141)) + list(range(181, 216)),
    'mt': list(range(0, 36)),
    'tn': list(range(0, 41)) + list(range(56, 181)),
    'tr': list(range(0, 66)) + list(range(106, 141)),
    'ad': list(range(0, 76))
}

processor = ExcelDataCreation('your_file_path_here.xlsx', sheets, rows)
processor.divide_excel_data('selected_output_path.xlsx', 'unselected_output_path.xlsx')


# Create new files for each sheet
for sheet in sheets:
    if sheet == 'pf':
        processor.create_new_file('pf', 0, ['pi'])
    elif sheet == 'pi':
        processor.create_new_file('pi', 0, ['pf', 'th'])
    elif sheet == 'th':
        processor.create_new_file('th', 0, ['pi'])
    elif sheet == 'es':
        processor.create_new_file('es', 0, ['gr', 'mt', 'tn', 'tr'])
    elif sheet == 'gr':
        processor.create_new_file('gr', 0, ['es', 'mt', 'tn', 'tr'])
    elif sheet == 'mt':
        processor.create_new_file('mt', 0, ['gr', 'es', 'tn', 'tr'])
    elif sheet == 'tn':
        processor.create_new_file('tn', 0, ['gr', 'mt', 'es', 'tr'])
    elif sheet == 'tr':
        processor.create_new_file('tr', 0, ['gr', 'mt', 'tn', 'es'])
        
###################################################################################################