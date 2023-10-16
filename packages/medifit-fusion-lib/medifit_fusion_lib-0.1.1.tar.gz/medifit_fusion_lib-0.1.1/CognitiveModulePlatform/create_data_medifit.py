# -*- coding: utf-8 -*-
"""
Created on Mon Feb  6 23:29:04 2023

@author: anton
"""

# import pandas as pd

# # Load the excel file
# file = 'file.xlsx'
# sheets = pd.read_excel(file, sheet_name=None)

# # Select specific rows from different sheets
# pf_rows = list(range(1,36)) + list(range(51, 86) + list(range(101, 136)) + list(range(151, 191)) + list(range(206, 281)))
# pf_selected = sheets['pf'].iloc[pf_rows]
# pf_unselected = sheets['pf'][~sheets['pf'].index.isin(pf_rows)]
# # ES1-7, GR1-7, MT1-7, TN 1-8, TR1-15


# th_rows = list(range(1, 101)) + list(range(146, 256) + list(range(301, 366)))
# th_selected = sheets['th'].iloc[th_rows]
# th_unselected = sheets['th'][~sheets['th'].index.isin(th_rows)]
# #GR1-20, TR0, TN1-22, ES1-13

# pi_rows = list(range(1, 36)) + list(range(51, 86))
# pi_selected = sheets['pi'].iloc[pi_rows]
# pi_unselected = sheets['pi'][~sheets['pi'].index.isin(pi_rows)]
# #GR1-7, TR1-7

# es_rows = list(range(1, 36)) + list(range(51, 111))
# es_selected = sheets['es'].iloc[es_rows]
# es_unselected = sheets['es'][~sheets['es'].index.isin(es_rows)]
# #Pf1-7, Th1-12

# gr_rows = list(range(1, 36)) + list(range(51, 141) + list(range(181, 216)))
# gr_selected = sheets['gr'].iloc[gr_rows]
# gr_unselected = sheets['gr'][~sheets['gr'].index.isin(gr_rows)]
# # Pf1-7, Th1-18, Pi1-7

# mt_rows = list(range(1, 36))
# mt_selected = sheets['mt'].iloc[mt_rows]
# mt_unselected = sheets['mt'][~sheets['mt'].index.isin(mt_rows)]
# #Pf1-7

# tn_rows = list(range(1, 41)) + list(range(56, 181))
# tn_selected = sheets['tn'].iloc[tn_rows]
# tn_unselected = sheets['tn'][~sheets['tn'].index.isin(tn_rows)]
# #Pf1-8, Th1-25

# tr_rows = list(range(1, 66)) + list(range(106, 141))
# tr_selected = sheets['tr'].iloc[tr_rows]
# tr_unselected = sheets['tr'][~sheets['tr'].index.isin(tr_rows)]
# #Pf1-13, Th0, Pi1-7

# ad_rows = list(range(1-76))
# ad_selected = sheets['ad'].iloc[ad_rows]
# ad_unselected = sheets['ad'][~sheets['ad'].index.isin(ad_rows)]
# #ThGR02

# # Concatenate the selected rows into a new data frame
# selected_data = pd.concat([pf_selected, th_selected, pi_selected, es_selected, gr_selected, mt_selected, tn_selected, tr_selected, ad_selected], ignore_index=True)
# unselected_data = pd.concat([pf_unselected, th_unselected, pi_unselected, es_unselected, gr_unselected, mt_unselected, tn_unselected, tr_unselected, ad_unselected], ignore_index=True)

# # Write the selected data to a new excel file
# selected_data.to_excel('selected_data.xlsx', index=False)
# unselected_data.to_excel('unselected_data.xlsx', index=False)

import pandas as pd

# Define the list of sheets
sheets = ['pf', 'th', 'pi', 'es', 'gr', 'mt', 'tn', 'tr', 'ad']

# Define the rows to select for each sheet
rows = {
    'pf': list(range(0,36)) + list(range(51, 86)) + list(range(101, 136)) + list(range(151, 191)) + list(range(206, 281)),
    'th': list(range(0, 101)) + list(range(146, 256)) + list(range(301, 366)),
    'pi': list(range(0, 36)) + list(range(51, 86)),
    'es': list(range(0, 36)) + list(range(51, 111)),
    'gr': list(range(0, 36)) + list(range(51, 141)) + list(range(181, 216)),
    'mt': list(range(0, 36)),
    'tn': list(range(0, 41)) + list(range(56, 181)),
    'tr': list(range(0, 66)) + list(range(106, 141)),
    'ad': list(range(0, 76))
}

# Load the Excel file
dfs = pd.read_excel(r"D:\Desktop\Folders\MEDIFIT_More\newer\Original Data\original_processed\FT-NIR_proc.xlsx", sheet_name=sheets)

# Load the excel file and select the rows for each sheet
selected_data = []
unselected_data = []
for sheet in sheets:
    df = dfs[sheet]
    selected = df.iloc[rows[sheet]].reset_index(drop=True)
    unselected = df[~df.index.isin(rows[sheet])].reset_index(drop=True)
    
    selected_data.append(selected)
    unselected_data.append(unselected)

# Merge the selected data into a single dataframe
selected_merged = pd.concat(selected_data, axis=0, keys=sheets, names=['sheet'])

# Write the merged selected data to a single Excel file with multiple sheets
selected_writer = pd.ExcelWriter(r'D:\Desktop\Folders\MEDIFIT_More\newer\Original Data\original_processed\selected_data1FTNIR.xlsx', engine='xlsxwriter')
for sheet in sheets:
    selected_merged.loc[sheet].to_excel(selected_writer, sheet_name=sheet, index=False, header=False)
selected_writer.save()

# Merge the unselected data into a single dataframe
unselected_merged = pd.concat(unselected_data, axis=0, keys=sheets, names=['sheet'])

# Write the merged unselected data to a single Excel file with multiple sheets
unselected_writer = pd.ExcelWriter(r'D:\Desktop\Folders\MEDIFIT_More\newer\Original Data\original_processed\unselected_data1FTNIR.xlsx', engine='xlsxwriter')
for sheet in sheets:
    unselected_merged.loc[sheet].to_excel(unselected_writer, sheet_name=sheet, index=False, header=False)
unselected_writer.save()

########################################THIS DOES EXACTLY THE SAME AS BEFORE BUT IN A FUNCTION###################

def process_excel_data(file_path, sheets, rows, selected_output_path, unselected_output_path):
    # Load the Excel file
    dfs = pd.read_excel(file_path, sheet_name=sheets)

    # Load the excel file and select the rows for each sheet
    selected_data = []
    unselected_data = []
    for sheet in sheets:
        df = dfs[sheet]
        selected = df.iloc[rows[sheet]].reset_index(drop=True)
        unselected = df[~df.index.isin(rows[sheet])].reset_index(drop=True)

        selected_data.append(selected)
        unselected_data.append(unselected)

    # Merge the selected data into a single dataframe
    selected_merged = pd.concat(selected_data, axis=0, keys=sheets, names=['sheet'])

    # Write the merged selected data to a single Excel file with multiple sheets
    selected_writer = pd.ExcelWriter(selected_output_path, engine='xlsxwriter')
    for sheet in sheets:
        selected_merged.loc[sheet].to_excel(selected_writer, sheet_name=sheet, index=False, header=False)
    selected_writer.save()

    # Merge the unselected data into a single dataframe
    unselected_merged = pd.concat(unselected_data, axis=0, keys=sheets, names=['sheet'])

    # Write the merged unselected data to a single Excel file with multiple sheets
    unselected_writer = pd.ExcelWriter(unselected_output_path, engine='xlsxwriter')
    for sheet in sheets:
        unselected_merged.loc[sheet].to_excel(unselected_writer, sheet_name=sheet, index=False, header=False)
    unselected_writer.save()

# Example of the function usage:
# process_excel_data(r'D:\Desktop\Folders\MEDIFIT_More\newer\Original Data\original_processed\Raman_proc.xlsx', 
                    #sheets, # So we need that rows list
                    #rows, # So we need to have defined the dict with rows that we select
                    #r'D:\Desktop\Folders\MEDIFIT_More\newer\Original Data\original_processed\selected_data1_Raman.xlsx', 
                    #r'D:\Desktop\Folders\MEDIFIT_More\newer\Original Data\original_processed\unselected_data1_Raman.xlsx'
                    #)


#########################################TRAINING - VALIDATION SPLITTING FROM EXCEL##############################
import openpyxl

# Load the existing workbook
wb = openpyxl.load_workbook(r"D:\Desktop\Folders\Honey\Original Data\Fusion1\Training.xlsx")

#### POLYFLORAL SCENARIO ####

# Get the sheets
pf_sheet = wb['pf']
pi_sheet = wb['pi']

# Create a new workbook
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Copy the rows from pf with a new first column added with value 0
for row in pf_sheet.iter_rows(values_only=True):
    new_row = [0] + list(row)
    new_sheet.append(new_row)

# Copy the rows from pi with the value 1 in the first column
for row in pi_sheet.iter_rows(values_only=True):
    new_row = [1] + list(row)
    new_sheet.append(new_row)

# Save the new workbook
new_wb.save(r"D:\Desktop\Folders\Honey\Original Data\Fusion1\Pine.xlsx")

#### THYME SCENARIO ####

# Get the sheets
th_sheet = wb['th']
pi_sheet = wb['pi']

# Create a new workbook
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Copy the rows from pf with a new first column added with value 0
for row in pf_sheet.iter_rows(values_only=True):
    new_row = [0] + list(row)
    new_sheet.append(new_row)

# Copy the rows from pi with the value 1 in the first column
for row in pi_sheet.iter_rows(values_only=True):
    new_row = [1] + list(row)
    new_sheet.append(new_row)

# Save the new workbook
new_wb.save(r"D:\Desktop\Folders\Honey\Original Data\Fusion1\Pine.xlsx")

#### PINE SCENARIO ####

# Get the sheets
pf_sheet = wb['pf']
pi_sheet = wb['pi']

# Create a new workbook
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Copy the rows from pf with a new first column added with value 0
for row in pf_sheet.iter_rows(values_only=True):
    new_row = [0] + list(row)
    new_sheet.append(new_row)

# Copy the rows from pi with the value 1 in the first column
for row in pi_sheet.iter_rows(values_only=True):
    new_row = [1] + list(row)
    new_sheet.append(new_row)

# Save the new workbook
new_wb.save(r"D:\Desktop\Folders\Honey\Original Data\Fusion1\Pine.xlsx")


import openpyxl

sheets = ['pf', 'pi', 'th', 'es', 'gr', 'mt', 'tn', 'tr']

def create_new_file(sheet, value_0, value_1):
    # Load the existing workbook
    wb = openpyxl.load_workbook(r"D:\Desktop\Folders\Honey\Original Data\Fusion1\Validation.xlsx")

    # Get the sheet
    sheet_to_copy = wb[sheet]

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Copy the rows from the sheet with a new first column added with value 0
    for row in sheet_to_copy.iter_rows(values_only=True):
        new_row = [value_0] + list(row)
        new_sheet.append(new_row)

    # Copy the rows from the other sheets with the value 1 in the first column
    for s in value_1:
        other_sheet = wb[s]
        for row in other_sheet.iter_rows(values_only=True):
            new_row = [1] + list(row)
            new_sheet.append(new_row)

    # Save the new workbook
    new_wb.save(f'D:/Desktop/Folders/Honey/Original Data/Fusion1/{sheet}_validation.xlsx')
    #new_wb.save(f'D:\Desktop\Folders\MEDIFIT_More\newer\Original Data\original_processed\FT-NIR\{sheet}_validation.xlsx')
    

# Create new files for each sheet
for sheet in sheets:
    if sheet == 'pf':
        create_new_file('pf', 0, ['pi'])
    elif sheet == 'pi':
        create_new_file('pi', 0, ['pf', 'th'])
    elif sheet == 'th':
        create_new_file('th', 0, ['pi'])
    elif sheet == 'es':
        create_new_file('es', 0, ['gr', 'mt', 'tn', 'tr'])
    elif sheet == 'gr':
        create_new_file('gr', 0, ['es', 'mt', 'tn', 'tr'])
    elif sheet == 'mt':
        create_new_file('mt', 0, ['gr', 'es', 'tn', 'tr'])
    elif sheet == 'tn':
        create_new_file('tn', 0, ['gr', 'mt', 'es', 'tr'])
    elif sheet == 'tr':
        create_new_file('tr', 0, ['gr', 'mt', 'tn', 'es'])

def create_new_file(sheet, value_0, value_1):
    # Load the existing workbook
    wb = openpyxl.load_workbook(r"D:\Desktop\Folders\Honey\Original Data\Fusion2\Training.xlsx")

    # Get the sheet
    sheet_to_copy = wb[sheet]

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Copy the rows from the sheet with a new first column added with value 0
    for row in sheet_to_copy.iter_rows(values_only=True):
        new_row = [value_0] + list(row)
        new_sheet.append(new_row)

    # Copy the rows from the other sheets with the value 1 in the first column
    for s in value_1:
        other_sheet = wb[s]
        for row in other_sheet.iter_rows(values_only=True):
            new_row = [1] + list(row)
            new_sheet.append(new_row)

    # Save the new workbook
    new_wb.save(f'D:/Desktop/Folders/Honey/Original Data/Fusion2/{sheet}_training.xlsx')

# Create new files for each sheet
for sheet in sheets:
    if sheet == 'pf':
        create_new_file('pf', 0, ['pi'])
    elif sheet == 'pi':
        create_new_file('pi', 0, ['pf', 'th'])
    elif sheet == 'th':
        create_new_file('th', 0, ['pi'])
    elif sheet == 'es':
        create_new_file('es', 0, ['gr', 'mt', 'tn', 'tr'])
    elif sheet == 'gr':
        create_new_file('gr', 0, ['es', 'mt', 'tn', 'tr'])
    elif sheet == 'mt':
        create_new_file('mt', 0, ['gr', 'es', 'tn', 'tr'])
    elif sheet == 'tn':
        create_new_file('tn', 0, ['gr', 'mt', 'es', 'tr'])
    elif sheet == 'tr':
        create_new_file('tr', 0, ['gr', 'mt', 'tn', 'es'])

def create_new_file(sheet, value_0, value_1):
    # Load the existing workbook
    wb = openpyxl.load_workbook(r"D:\Desktop\Folders\Honey\Original Data\Fusion2\Validation.xlsx")

    # Get the sheet
    sheet_to_copy = wb[sheet]

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Copy the rows from the sheet with a new first column added with value 0
    for row in sheet_to_copy.iter_rows(values_only=True):
        new_row = [value_0] + list(row)
        new_sheet.append(new_row)

    # Copy the rows from the other sheets with the value 1 in the first column
    for s in value_1:
        other_sheet = wb[s]
        for row in other_sheet.iter_rows(values_only=True):
            new_row = [1] + list(row)
            new_sheet.append(new_row)

    # Save the new workbook
    new_wb.save(f'D:/Desktop/Folders/Honey/Original Data/Fusion2/{sheet}_validation.xlsx')

# Create new files for each sheet
for sheet in sheets:
    if sheet == 'pf':
        create_new_file('pf', 0, ['pi'])
    elif sheet == 'pi':
        create_new_file('pi', 0, ['pf', 'th'])
    elif sheet == 'th':
        create_new_file('th', 0, ['pi'])
    elif sheet == 'es':
        create_new_file('es', 0, ['gr', 'mt', 'tn', 'tr'])
    elif sheet == 'gr':
        create_new_file('gr', 0, ['es', 'mt', 'tn', 'tr'])
    elif sheet == 'mt':
        create_new_file('mt', 0, ['gr', 'es', 'tn', 'tr'])
    elif sheet == 'tn':
        create_new_file('tn', 0, ['gr', 'mt', 'es', 'tr'])
    elif sheet == 'tr':
        create_new_file('tr', 0, ['gr', 'mt', 'tn', 'es'])

def create_new_file(sheet, value_0, value_1):
    # Load the existing workbook
    wb = openpyxl.load_workbook(r"D:\Desktop\Folders\Honey\Original Data\Fusion3\Training.xlsx")

    # Get the sheet
    sheet_to_copy = wb[sheet]

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Copy the rows from the sheet with a new first column added with value 0
    for row in sheet_to_copy.iter_rows(values_only=True):
        new_row = [value_0] + list(row)
        new_sheet.append(new_row)

    # Copy the rows from the other sheets with the value 1 in the first column
    for s in value_1:
        other_sheet = wb[s]
        for row in other_sheet.iter_rows(values_only=True):
            new_row = [1] + list(row)
            new_sheet.append(new_row)

    # Save the new workbook
    new_wb.save(f'D:/Desktop/Folders/Honey/Original Data/Fusion3/{sheet}_training.xlsx')

# Create new files for each sheet
for sheet in sheets:
    if sheet == 'pf':
        create_new_file('pf', 0, ['pi'])
    elif sheet == 'pi':
        create_new_file('pi', 0, ['pf', 'th'])
    elif sheet == 'th':
        create_new_file('th', 0, ['pi'])
    elif sheet == 'es':
        create_new_file('es', 0, ['gr', 'mt', 'tn', 'tr'])
    elif sheet == 'gr':
        create_new_file('gr', 0, ['es', 'mt', 'tn', 'tr'])
    elif sheet == 'mt':
        create_new_file('mt', 0, ['gr', 'es', 'tn', 'tr'])
    elif sheet == 'tn':
        create_new_file('tn', 0, ['gr', 'mt', 'es', 'tr'])
    elif sheet == 'tr':
        create_new_file('tr', 0, ['gr', 'mt', 'tn', 'es'])

def create_new_file(sheet, value_0, value_1):
    # Load the existing workbook
    wb = openpyxl.load_workbook(r"D:\Desktop\Folders\Honey\Original Data\Fusion3\Validation.xlsx")

    # Get the sheet
    sheet_to_copy = wb[sheet]

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Copy the rows from the sheet with a new first column added with value 0
    for row in sheet_to_copy.iter_rows(values_only=True):
        new_row = [value_0] + list(row)
        new_sheet.append(new_row)

    # Copy the rows from the other sheets with the value 1 in the first column
    for s in value_1:
        other_sheet = wb[s]
        for row in other_sheet.iter_rows(values_only=True):
            new_row = [1] + list(row)
            new_sheet.append(new_row)

    # Save the new workbook
    new_wb.save(f'D:/Desktop/Folders/Honey/Original Data/Fusion3/{sheet}_validation.xlsx')

# Create new files for each sheet
for sheet in sheets:
    if sheet == 'pf':
        create_new_file('pf', 0, ['pi'])
    elif sheet == 'pi':
        create_new_file('pi', 0, ['pf', 'th'])
    elif sheet == 'th':
        create_new_file('th', 0, ['pi'])
    elif sheet == 'es':
        create_new_file('es', 0, ['gr', 'mt', 'tn', 'tr'])
    elif sheet == 'gr':
        create_new_file('gr', 0, ['es', 'mt', 'tn', 'tr'])
    elif sheet == 'mt':
        create_new_file('mt', 0, ['gr', 'es', 'tn', 'tr'])
    elif sheet == 'tn':
        create_new_file('tn', 0, ['gr', 'mt', 'es', 'tr'])
    elif sheet == 'tr':
        create_new_file('tr', 0, ['gr', 'mt', 'tn', 'es'])
        
###################################################################################################